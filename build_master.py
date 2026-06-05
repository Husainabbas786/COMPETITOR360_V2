# -*- coding: utf-8 -*-
"""
Competitor 360 — Master data builder.
Single source of truth -> Competitor360_Master.xlsx (human-editable, formula-driven)
                        -> master-data.json (the tool reads this).

Both artefacts are generated from ONE in-memory data model so they can never drift.
Every figure carries a confidence flag (confirmed / assumed / missing) and a source note.
Totals (BASE, ALL-IN, take-home) are written as Excel FORMULAS, never hardcoded.
"""

import json
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Confidence + styling
# ---------------------------------------------------------------------------
CONFIRMED, ASSUMED, MISSING, NA = "confirmed", "assumed", "missing", "na"

LEGEND = {
    CONFIRMED: ("✅ Confirmed", "official source"),
    ASSUMED:   ("🟡 Assumed",   "estimate / placeholder — overwrite when real data lands"),
    MISSING:   ("⬜ Missing",   "no data yet"),
    NA:        ("—  N/A",       "not applicable for this zone"),
}

FILL = {
    CONFIRMED: PatternFill("solid", fgColor="C6EFCE"),  # green
    ASSUMED:   PatternFill("solid", fgColor="FFEB9C"),  # yellow
    MISSING:   PatternFill("solid", fgColor="D9D9D9"),  # grey
    NA:        PatternFill("solid", fgColor="F2F2F2"),  # very light grey
}
FONTCOL = {
    CONFIRMED: "006100",
    ASSUMED:   "9C6500",
    MISSING:   "808080",
    NA:        "BFBFBF",
}
FORMULA_FILL = PatternFill("solid", fgColor="DCE6F1")   # light blue = derived
TEAL_FILL    = PatternFill("solid", fgColor="0F8B8D")   # Meydan accent
INK = "1A1A1A"
ARIAL = "Arial"

def font(size=10, bold=False, color=INK, italic=False):
    return Font(name=ARIAL, size=size, bold=bold, color=color, italic=italic)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")

ZONES = ["Meydan", "IFZA", "RAKEZ", "Ajman", "DWTC"]

# ---------------------------------------------------------------------------
# Cell helpers — produce a uniform dict for every data point
#   kind: num | incl | na | missing
#   numeric: the value that enters Excel/JSON computations (incl/na/blank -> 0)
# ---------------------------------------------------------------------------
def num(v, conf, note=None):
    return {"kind": "num", "value": v, "numeric": v, "display": f"{v:,.0f}",
            "confidence": conf, "note": note}

def incl(conf, note=None):
    return {"kind": "incl", "value": None, "numeric": 0, "display": "incl",
            "confidence": conf, "note": note}

def na(note=None):
    return {"kind": "na", "value": None, "numeric": 0, "display": "—",
            "confidence": NA, "note": note}

def miss(value=None, note=None):
    # value optional (e.g. DWTC status change = 0 but unconfirmed)
    disp = f"{value:,.0f}" if value is not None else ""
    return {"kind": "missing", "value": value, "numeric": value or 0, "display": disp,
            "confidence": MISSING, "note": note}

def pct(v, conf, note=None):
    return {"kind": "pct", "value": v, "numeric": v, "display": f"{v*100:.0f}%",
            "confidence": conf, "note": note}

def txt(v, conf, note=None):
    return {"kind": "text", "value": v, "numeric": None, "display": v,
            "confidence": conf, "note": note}

# ===========================================================================
# DATA MODEL
# ===========================================================================

# --- B2C component matrix --------------------------------------------------
# Each row: (label, {zone: cell}). Order matters for formulas.
B2C_ROWS = [
    ("Licence (0-visa, 3 activities)", {
        "Meydan": num(12125, CONFIRMED),
        "IFZA":   num(12900, ASSUMED, "Contact says 11,900 — confirm IFZA 11,900 vs 12,900"),
        "RAKEZ":  num(6010, CONFIRMED),
        "Ajman":  num(4888, CONFIRMED),
        "DWTC":   num(12000, ASSUMED),
    }),
    ("Registration / admin", {
        "Meydan": incl(CONFIRMED),
        "IFZA":   incl(CONFIRMED),
        "RAKEZ":  incl(CONFIRMED),
        "Ajman":  incl(CONFIRMED),
        "DWTC":   num(500, ASSUMED),
    }),
    ("Shared desk / co-working", {
        "Meydan": num(375, CONFIRMED),
        "IFZA":   incl(ASSUMED),
        "RAKEZ":  incl(CONFIRMED),
        "Ajman":  incl(CONFIRMED, "included via lease"),
        "DWTC":   incl(MISSING, "desk cost unknown — confirm DWTC desk"),
    }),
    ("Establishment card", {
        "Meydan": num(2000, CONFIRMED),
        "IFZA":   num(2000, CONFIRMED),
        "RAKEZ":  incl(CONFIRMED),
        "Ajman":  incl(CONFIRMED),
        "DWTC":   num(2300, ASSUMED),
    }),
    ("Visa allocation (per visa)", {
        "Meydan": num(1850, CONFIRMED),
        "IFZA":   num(2000, CONFIRMED, "licence-tier step"),
        "RAKEZ":  incl(CONFIRMED),
        "Ajman":  incl(CONFIRMED),
        "DWTC":   num(2000, ASSUMED, "tier step"),
    }),
    ("Residence visa – investor (per visa)", {
        "Meydan": num(4000, CONFIRMED),
        "IFZA":   num(1000, CONFIRMED, "Visa FREE-for-life (0) + 1,000 title deed"),
        "RAKEZ":  incl(CONFIRMED),
        "Ajman":  incl(CONFIRMED),
        "DWTC":   num(3020, ASSUMED),
    }),
    ("Medical (per visa)", {
        "Meydan": num(2000, ASSUMED),
        "IFZA":   num(700, ASSUMED, "~700 estimate"),
        "RAKEZ":  incl(CONFIRMED, "+40 allocation fee"),
        "Ajman":  incl(CONFIRMED),
        "DWTC":   miss(note="DWTC medical cost unknown"),
    }),
    ("Emirates ID (per visa)", {
        "Meydan": num(750, ASSUMED),
        "IFZA":   num(400, ASSUMED, "~400 estimate"),
        "RAKEZ":  incl(CONFIRMED, "+30 delivery fee"),
        "Ajman":  incl(CONFIRMED),
        "DWTC":   miss(note="DWTC Emirates ID cost unknown"),
    }),
    ("Status change (inside UAE)", {
        "Meydan": num(1500, CONFIRMED),
        "IFZA":   num(1600, CONFIRMED),
        "RAKEZ":  incl(CONFIRMED),
        "Ajman":  incl(CONFIRMED),
        "DWTC":   miss(0, "assumed 0 — confirm DWTC status change"),
    }),
    ("Visa bundle increment (bundled zones)", {
        "Meydan": na(),
        "IFZA":   na(),
        "RAKEZ":  num(8000, CONFIRMED, "RAKEZ bundles all visa components into this increment"),
        "Ajman":  num(5912, CONFIRMED, "Ajman bundles all visa components into this increment"),
        "DWTC":   na(),
    }),
]

# Expected headline figures (used for manual formula verification)
EXPECT_BASE  = {"Meydan": 12500, "IFZA": 12900, "RAKEZ": 6010, "Ajman": 4888, "DWTC": 12500}
EXPECT_ALLIN = {"Meydan": 24600, "IFZA": 20600, "RAKEZ": 14010, "Ajman": 10800, "DWTC": 19820}

B2C_NOTE = ("BASE = Licence + Registration + Shared Desk (0 visa). "
            "ALL-IN = sum of all components for 1 investor visa, Year 1. "
            "'incl' / '—' / blank cells contribute 0 to the sums. "
            "Meydan ALL-IN = 24,100 if the bundled Medical+EID Assistance (2,250) is taken "
            "instead of itemised Medical 2,000 + EID 750 (2,750).")

# --- B2B commission --------------------------------------------------------
# Per-zone summary rows (NOT per-tier): model, what the % applies to, base
# amount, renewal. The per-tier RATES live in the uniform B2B_TIERS below.
B2B = {
    "model": {
        "Meydan": txt("5 tiers (quarterly); CP picks one-time OR recurring", CONFIRMED),
        "IFZA":   txt("Plan A (%) or Plan B (one-time)", CONFIRMED),
        "RAKEZ":  txt("Flat", CONFIRMED),
        "Ajman":  txt("Tiered (yearly companies)", CONFIRMED),
        "DWTC":   txt("Flat", CONFIRMED),
    },
    "base_applies": {
        "Meydan": txt("Full package", ASSUMED),
        "IFZA":   txt("Standard licence price", CONFIRMED),
        "RAKEZ":  txt("Licence only", ASSUMED),
        "Ajman":  txt("All-inclusive package", CONFIRMED),
        "DWTC":   txt("Unknown", ASSUMED),
    },
    "base_amount": {
        "Meydan": num(12500, ASSUMED),
        "IFZA":   num(14900, CONFIRMED),
        "RAKEZ":  num(6010, ASSUMED),
        "Ajman":  num(10800, CONFIRMED),
        "DWTC":   num(14000, ASSUMED),
    },
    "renewal": {
        "Meydan": pct(0.40, CONFIRMED, "renewal rate"),
        "IFZA":   na("Plan B has no renewal"),
        "RAKEZ":  miss(note="RAKEZ renewal unknown"),
        "Ajman":  pct(0.35, CONFIRMED, "flat"),
        "DWTC":   miss(note="DWTC renewal unknown"),
    },
}
B2B_ROW_ORDER = [
    ("Model", "model"),
    ("Commission BASE (% applies to)", "base_applies"),
    ("Base amount (1-visa, AED)", "base_amount"),
    ("Renewal rate", "renewal"),
]
EXPECT_TAKEHOME = {"Meydan": 3750, "IFZA": 5960, "RAKEZ": 3005, "Ajman": 3780, "DWTC": 2800}

# --- B2B UNIFORM TIER SCHEMA ----------------------------------------------
# One shape for every zone: ordered entry -> top, each tier = rank, label,
# one_time, recurring (decimal or None), per-rate confidence + tier note.
# Tiered zones use their real ladder; non-tiered zones collapse to ONE constant
# tier. Rates are taken verbatim from the master — nothing invented or changed.
CONF_RANK_PY = {"confirmed": 1, "assumed": 2, "missing": 3, "na": 0}
def _worst(cs):
    cs = [c for c in cs if c]
    return max(cs, key=lambda c: CONF_RANK_PY.get(c, 0)) if cs else MISSING

def tier(rank, label, one_time, recurring, ot_conf, rec_conf, note=None):
    present = []
    if one_time is not None: present.append(ot_conf)
    if recurring is not None: present.append(rec_conf)
    return {
        "rank": rank, "label": label,
        "one_time": one_time, "recurring": recurring,
        "one_time_conf": ot_conf if one_time is not None else MISSING,
        "recurring_conf": rec_conf if recurring is not None else MISSING,
        "confidence": _worst(present),
        "note": note,
    }

B2B_TIERS = {
    # Meydan — 5 quarterly new-sales tiers (one-time / recurring), all confirmed.
    "Meydan": [
        tier(1, "Bronze · 0–20/qtr",     0.30, 0.20, CONFIRMED, CONFIRMED),
        tier(2, "Silver · 21–30/qtr",    0.35, 0.25, CONFIRMED, CONFIRMED),
        tier(3, "Gold · 31–75/qtr",      0.40, 0.30, CONFIRMED, CONFIRMED),
        tier(4, "Platinum · 76–120/qtr", 0.40, 0.40, CONFIRMED, CONFIRMED),
        tier(5, "Platinum+ · 121+/qtr",  0.50, 0.50, CONFIRMED, CONFIRMED, "50% new, 40% on renewal"),
    ],
    # Ajman — 5 yearly-company tiers; recurring/renewal flat 35% at every tier.
    "Ajman": [
        tier(1, "≤10 cos/yr",  0.35, 0.35, CONFIRMED, CONFIRMED),
        tier(2, "≤20 cos/yr",  0.40, 0.35, CONFIRMED, CONFIRMED),
        tier(3, "≤30 cos/yr",  0.45, 0.35, CONFIRMED, CONFIRMED),
        tier(4, "≤350 cos/yr", 0.50, 0.35, CONFIRMED, CONFIRMED),
        tier(5, "350+ cos/yr", 0.55, 0.35, CONFIRMED, CONFIRMED, "capped 52% on a 1-visa deal; recurring/renewal flat 35%"),
    ],
    # IFZA — not tiered: a single constant tier (Plan B one-time / Plan A recurring).
    "IFZA": [
        tier(1, "Flat (Plan A / Plan B)", 0.40, 0.20, CONFIRMED, ASSUMED, "Plan B 40% one-time; Plan A recurring 20–25% (assumed)"),
    ],
    # RAKEZ — flat; recurring not disclosed.
    "RAKEZ": [
        tier(1, "Flat", 0.50, None, CONFIRMED, MISSING, "Single source; recurring not disclosed"),
    ],
    # DWTC — flat; recurring not disclosed.
    "DWTC": [
        tier(1, "Flat", 0.20, None, CONFIRMED, MISSING, "Recurring not disclosed"),
    ],
}
def tier_entry(zone): return B2B_TIERS[zone][0]
def tier_top(zone):   return B2B_TIERS[zone][-1]

# Caveat the tool surfaces near the level control (kept in the master, not hardcoded in the UI).
B2B_LEVEL_NOTE = ("Level = each zone's own ladder position (Meydan = sales/quarter, "
                  "Ajman = companies/year), not an identical volume.")

# Plain-language meaning of each commitment level (shown beneath the control).
B2B_LEVEL_DESCRIPTIONS = {
    "low":  "Entry tier of each zone's commission ladder: a partner with little or no business volume.",
    "mid":  "Middle tier of each zone's ladder: a partner with moderate business volume.",
    "high": "Top tier of each zone's ladder: a partner at the highest qualifying volume.",
}

# --- Multi-year & Year-2 ---------------------------------------------------
MULTIYEAR_DISCOUNT = {
    "Meydan": txt("2yr 10% / 3yr+ 15%", ASSUMED),
    "IFZA":   txt("2yr 15% / 3yr 20% / 5yr 30%", CONFIRMED),
    "RAKEZ":  txt("Explicit non-linear table (see below)", CONFIRMED),
    "Ajman":  txt("Annual model — no multi-year", CONFIRMED),
    "DWTC":   miss(note="DWTC multi-year unknown"),
}
RAKEZ_MY_TABLE = {  # Biz One package, AED cumulative
    "1y": num(14010, CONFIRMED), "2y": num(26620, CONFIRMED), "3y": num(37830, CONFIRMED),
    "4y": num(50440, CONFIRMED), "5y": num(59560, CONFIRMED), "6y": num(67260, CONFIRMED),
    "10y": num(105100, CONFIRMED),
}
YEAR2 = {
    "Meydan": num(14700, ASSUMED, "~14,700 — assumes visa NOT re-charged in Year 2"),
    "IFZA":   num(17100, ASSUMED, "~17,100"),
    "RAKEZ":  num(14010, CONFIRMED),
    "Ajman":  num(9900, CONFIRMED),
    "DWTC":   miss(note="DWTC Year-2 unknown"),
}
YEAR2_KEY = ("Year-2 powers the flagship inference: 'Year-1 Meydan costly, Year-2 Meydan cheaper.' "
             "It rests on the assumption that the residence visa (valid ~2 yrs) is NOT re-charged in "
             "Year 2. CONFIRM with Jai — this is the lever.")

# --- Sources ---------------------------------------------------------------
SOURCES = [
    # Area, Zone, Source, Confidence, Note
    ["B2C licence/visa fees", "Meydan", "Meydan official quote", CONFIRMED, "Baseline zone"],
    ["B2C medical/EID",       "Meydan", "Internal estimate",     ASSUMED,  "Confirm real cost"],
    ["B2C licence",           "IFZA",   "IFZA quote / contact",  ASSUMED,  "11,900 vs 12,900 open"],
    ["B2C medical/EID",       "IFZA",   "Internal estimate",     ASSUMED,  "Confirm real cost"],
    ["B2C all fees",          "RAKEZ",  "RAKEZ official (Biz One)", CONFIRMED, "Bundled visa increment"],
    ["B2C all fees",          "Ajman",  "Ajman official",        CONFIRMED, "Bundled visa increment"],
    ["B2C fees",              "DWTC",   "Partial / estimate",    ASSUMED,  "Desk/medical/EID/status missing"],
    ["B2B commission base",   "Meydan", "To confirm (Akshay)",   ASSUMED,  "Full package assumed"],
    ["B2B commission",        "IFZA",   "IFZA partner terms",    CONFIRMED, "Plan A/B"],
    ["B2B commission base",   "RAKEZ",  "To confirm (Akshay)",   ASSUMED,  "Licence-only assumed"],
    ["B2B commission",        "Ajman",  "Ajman partner terms",   CONFIRMED, "Tiered yearly"],
    ["B2B commission base",   "DWTC",   "To confirm (Akshay)",   ASSUMED,  "Base unknown"],
    ["Multi-year",            "RAKEZ",  "RAKEZ Biz One schedule", CONFIRMED, "Explicit table"],
    ["Year-2 visa-recharge",  "All",    "To confirm (Jai)",      ASSUMED,  "Flagship lever"],
]

# --- Assumptions & gaps (from 'GAPS TO SWAP IN' + every 🟡/⬜ cell) ---------
GAPS = [
    # Item, Current placeholder, Confidence, Owner, Action
    ["Commission bases (Meydan, RAKEZ, DWTC)", "Full package / Licence only / Unknown", ASSUMED, "Akshay", "Get the % base each zone applies commission to"],
    ["Multi-year (Meydan, DWTC)", "Meydan 10/15% est; DWTC missing", ASSUMED, "Akshay", "Get real multi-year discount schedules"],
    ["RAKEZ one-time vs recurring", "Recurring missing", MISSING, "Akshay", "Get RAKEZ recurring commission"],
    ["IFZA 20-vs-25% & 11,900-vs-12,900", "Plan A 20%; licence 12,900", ASSUMED, "Akshay", "Resolve both IFZA ranges"],
    ["Meydan & IFZA medical/EID real cost", "Meydan 2,000/750; IFZA 700/400 est", ASSUMED, "Akshay", "Get actual medical + Emirates ID costs"],
    ["DWTC desk/medical/EID/status/renewal", "Mostly missing", MISSING, "Akshay", "Get full DWTC component breakdown"],
    ["Live offers (all zones)", "None loaded", MISSING, "Akshay", "Get current promotional offers"],
    ["Year-2 visa-recharge logic", "Assumes visa not re-charged Y2", ASSUMED, "Jai", "Confirm visa validity / Year-2 recharge — THE LEVER"],
]

# ===========================================================================
# Workbook builders
# ===========================================================================
wb = Workbook()

def style_cell(c, cell_data, formula=False):
    """Apply value, fill, font and source comment to an openpyxl cell."""
    conf = cell_data["confidence"]
    c.font = font(color=FONTCOL.get(conf, INK))
    c.fill = FILL.get(conf, PatternFill())
    c.border = BORDER
    kind = cell_data["kind"]
    if kind in ("num",):
        c.value = cell_data["value"]; c.number_format = "#,##0"; c.alignment = RIGHT
    elif kind == "pct":
        c.value = cell_data["value"]; c.number_format = "0%"; c.alignment = CENTER
    elif kind == "missing":
        if cell_data["value"] is not None:
            c.value = cell_data["value"]; c.number_format = "#,##0"; c.alignment = RIGHT
        else:
            c.value = ""; c.alignment = CENTER
    else:  # incl, na, text
        c.value = cell_data["display"]; c.alignment = CENTER if kind in ("incl", "na") else LEFT
    if cell_data.get("note"):
        c.comment = Comment(cell_data["note"], "Competitor360")
        c.comment.width = 220; c.comment.height = 110

def header_row(ws, row, labels, label0="", teal_meydan=True):
    ws.cell(row=row, column=1, value=label0)
    cc = ws.cell(row=row, column=1); cc.font = font(bold=True, color="FFFFFF"); \
        cc.fill = PatternFill("solid", fgColor="404040"); cc.alignment = LEFT; cc.border = BORDER
    for i, lab in enumerate(labels):
        col = 2 + i
        c = ws.cell(row=row, column=col, value=lab)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = TEAL_FILL if (teal_meydan and lab == "Meydan") else PatternFill("solid", fgColor="404040")
        c.alignment = CENTER; c.border = BORDER

def title_block(ws, title, subtitle=None):
    ws.cell(row=1, column=1, value=title).font = font(16, bold=True, color="0F8B8D")
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = font(9, italic=True, color="595959")

# ---- README ---------------------------------------------------------------
def build_readme():
    ws = wb.active; ws.title = "README"
    ws.sheet_view.showGridLines = False
    title_block(ws, "Competitor 360 — Master Data",
                "Free-zone pricing (B2C) & commission (B2B) benchmark · Meydan = baseline vs IFZA, RAKEZ, Ajman, DWTC")
    r = 4
    lines = [
        ("This workbook is the SINGLE SOURCE OF TRUTH.", True),
        ("It generates master-data.json, which the comparison tool reads. Edit here, re-export, never edit the JSON by hand.", False),
        ("", False),
        ("Every figure is colour-coded by confidence. All totals are Excel FORMULAS — never type over a blue cell.", True),
        ("", False),
    ]
    for text, bold in lines:
        ws.cell(row=r, column=1, value=text).font = font(10, bold=bold); r += 1

    ws.cell(row=r, column=1, value="Confidence legend").font = font(11, bold=True); r += 1
    for key in (CONFIRMED, ASSUMED, MISSING, NA):
        lab, desc = LEGEND[key]
        c = ws.cell(row=r, column=1, value="  "); c.fill = FILL[key]; c.border = BORDER; c.font = font()
        ws.cell(row=r, column=2, value=lab).font = font(10, bold=True, color=FONTCOL[key])
        ws.cell(row=r, column=3, value=desc).font = font(10)
        r += 1
    c = ws.cell(row=r, column=1, value="  "); c.fill = FORMULA_FILL; c.border = BORDER; c.font = font()
    ws.cell(row=r, column=2, value="🔵 Derived").font = font(10, bold=True, color="1F4E79")
    ws.cell(row=r, column=3, value="computed by Excel formula (BASE, ALL-IN, take-home)").font = font(10)
    r += 2

    ws.cell(row=r, column=1, value="Sheets").font = font(11, bold=True); r += 1
    for name, desc in [
        ("B2C – Pricing", "Component matrix per visa. BASE = Licence+Registration+Desk; ALL-IN = SUM of all components."),
        ("B2B – Commission", "Model, base, base amount, rate ladders. Take-home = entry one-time rate × base amount."),
        ("Multi-year & Year-2", "Multi-year discounts, RAKEZ schedule, Year-2 cost (the flagship lever)."),
        ("Sources", "Where each figure came from + confidence."),
        ("Assumptions & Gaps", "Every placeholder to swap in, with owner."),
    ]:
        ws.cell(row=r, column=1, value=name).font = font(10, bold=True)
        ws.cell(row=r, column=3, value=desc).font = font(10)
        r += 1
    r += 1
    ws.cell(row=r, column=1,
            value="Key insight: take-home = rate × base. A lower rate on a bigger base can beat a higher rate "
                  "on a small one — Meydan 30%×12,500 = 3,750 beats RAKEZ 50%×6,010 = 3,005.").font = \
        font(10, italic=True, color="0F8B8D")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 90

# ---- B2C ------------------------------------------------------------------
def build_b2c():
    ws = wb.create_sheet("B2C – Pricing")
    ws.sheet_view.showGridLines = False
    title_block(ws, "B2C — Pricing (AED, VAT-incl)", "Components for one investor visa, Year 1")
    hr = 4
    header_row(ws, hr, ZONES, "Component")
    first_data = hr + 1
    row = first_data
    row_index = {}
    for label, cells in B2C_ROWS:
        rc = ws.cell(row=row, column=1, value=label)
        rc.font = font(10, bold=True); rc.alignment = LEFT; rc.border = BORDER
        rc.fill = PatternFill("solid", fgColor="F2F2F2")
        for i, z in enumerate(ZONES):
            style_cell(ws.cell(row=row, column=2 + i), cells[z])
        row_index[label] = row
        row += 1
    last_data = row - 1
    licence_r = row_index["Licence (0-visa, 3 activities)"]
    desk_r = row_index["Shared desk / co-working"]   # Licence, Registration, Desk are contiguous

    # BASE formula row
    base_row = row
    bc = ws.cell(row=base_row, column=1, value="BASE = Licence + Registration + Shared Desk")
    bc.font = font(10, bold=True, color="1F4E79"); bc.alignment = LEFT; bc.border = BORDER
    bc.fill = FORMULA_FILL
    for i in range(len(ZONES)):
        col = get_column_letter(2 + i)
        c = ws.cell(row=base_row, column=2 + i,
                    value=f"=SUM({col}{licence_r}:{col}{desk_r})")
        c.font = font(10, bold=True, color="1F4E79"); c.number_format = "#,##0"
        c.alignment = RIGHT; c.fill = FORMULA_FILL; c.border = BORDER
    row += 1

    # ALL-IN formula row
    allin_row = row
    ac = ws.cell(row=allin_row, column=1, value="ALL-IN = 1 investor visa, Year 1")
    ac.font = font(10, bold=True, color="1F4E79"); ac.alignment = LEFT; ac.border = BORDER
    ac.fill = FORMULA_FILL
    for i in range(len(ZONES)):
        col = get_column_letter(2 + i)
        c = ws.cell(row=allin_row, column=2 + i,
                    value=f"=SUM({col}{first_data}:{col}{last_data})")
        c.font = font(10, bold=True, color="1F4E79"); c.number_format = "#,##0"
        c.alignment = RIGHT; c.fill = FORMULA_FILL; c.border = BORDER
    row += 2

    ws.cell(row=row, column=1, value="Notes:").font = font(9, bold=True)
    ws.cell(row=row + 1, column=1, value=B2C_NOTE).font = font(9, italic=True, color="595959")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 3, end_column=6)
    ws.cell(row=row + 1, column=1).alignment = LEFT

    ws.column_dimensions["A"].width = 38
    for i in range(len(ZONES)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 14
    ws.freeze_panes = "B5"
    return {"base_row": base_row, "allin_row": allin_row}

# ---- B2B ------------------------------------------------------------------
def build_b2b():
    ws = wb.create_sheet("B2B – Commission")
    ws.sheet_view.showGridLines = False
    title_block(ws, "B2B — Commission", "Take-home = entry one-time rate × base amount")
    hr = 4
    header_row(ws, hr, ZONES, "Item")
    row = hr + 1
    row_index = {}
    for label, key in B2B_ROW_ORDER:
        rc = ws.cell(row=row, column=1, value=label)
        rc.font = font(10, bold=True); rc.alignment = LEFT; rc.border = BORDER
        rc.fill = PatternFill("solid", fgColor="F2F2F2")
        for i, z in enumerate(ZONES):
            style_cell(ws.cell(row=row, column=2 + i), B2B[key][z])
        row_index[key] = row
        row += 1

    # Entry / top one-time rate, projected from the uniform tier schema (one source).
    er_row = row
    rc = ws.cell(row=er_row, column=1, value="One-time — entry rate (tier rank 1)")
    rc.font = font(10, bold=True); rc.alignment = LEFT; rc.border = BORDER; rc.fill = PatternFill("solid", fgColor="F2F2F2")
    for i, z in enumerate(ZONES):
        style_cell(ws.cell(row=er_row, column=2 + i), pct(tier_entry(z)["one_time"], tier_entry(z)["one_time_conf"]))
    row += 1
    tr_row = row
    rc = ws.cell(row=tr_row, column=1, value="One-time — top rate (tier rank N)")
    rc.font = font(10, bold=True); rc.alignment = LEFT; rc.border = BORDER; rc.fill = PatternFill("solid", fgColor="F2F2F2")
    for i, z in enumerate(ZONES):
        style_cell(ws.cell(row=tr_row, column=2 + i), pct(tier_top(z)["one_time"], tier_top(z)["one_time_conf"]))
    row += 1

    base_r = row_index["base_amount"]
    # Take-home formula row = entry one-time rate × base amount
    th_row = row
    tc = ws.cell(row=th_row, column=1, value="Take-home (entry, 1-visa, one-time) = rate × base")
    tc.font = font(10, bold=True, color="1F4E79"); tc.alignment = LEFT; tc.border = BORDER
    tc.fill = FORMULA_FILL
    for i in range(len(ZONES)):
        col = get_column_letter(2 + i)
        c = ws.cell(row=th_row, column=2 + i, value=f"={col}{er_row}*{col}{base_r}")
        c.font = font(10, bold=True, color="1F4E79"); c.number_format = "#,##0"
        c.alignment = RIGHT; c.fill = FORMULA_FILL; c.border = BORDER
    row += 2

    # Uniform tier schema (one shape for every zone)
    ws.cell(row=row, column=1, value="Uniform commission tiers — entry → top (orange = ceiling)").font = font(12, bold=True, color="0F8B8D")
    row += 1
    tier_headers = ["Zone", "Rank", "Tier", "One-time", "Recurring", "Note"]
    for j, h in enumerate(tier_headers):
        c = ws.cell(row=row, column=1 + j, value=h)
        c.font = font(10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="595959")
        c.alignment = LEFT if j in (0, 2, 5) else CENTER; c.border = BORDER
    row += 1
    ORANGE = PatternFill("solid", fgColor="F4B183")
    for z in ZONES:
        zone_tiers = B2B_TIERS[z]
        for t in zone_tiers:
            zc = ws.cell(row=row, column=1, value=z); zc.font = font(10, bold=True); zc.border = BORDER
            rk = ws.cell(row=row, column=2, value=t["rank"]); rk.font = font(10); rk.alignment = CENTER; rk.border = BORDER
            lb = ws.cell(row=row, column=3, value=t["label"]); lb.font = font(10); lb.border = BORDER
            ot = pct(t["one_time"], t["one_time_conf"]) if t["one_time"] is not None else miss(note="not disclosed")
            style_cell(ws.cell(row=row, column=4), ot)
            if t is zone_tiers[-1] and t["one_time"] is not None:  # ceiling marker on top tier
                ws.cell(row=row, column=4).fill = ORANGE; ws.cell(row=row, column=4).font = font(10, bold=True, color="843C0C")
            rec = pct(t["recurring"], t["recurring_conf"]) if t["recurring"] is not None else miss(note="not disclosed")
            style_cell(ws.cell(row=row, column=5), rec)
            style_cell(ws.cell(row=row, column=6), txt(t["note"] or "", t["confidence"]))
            row += 1
    row += 1
    ws.cell(row=row, column=1,
            value="Key: take-home = rate × base. A lower rate on a bigger base can beat a higher rate "
                  "on a small one — Meydan 30%×12,500 = 3,750 beats RAKEZ 50%×6,010 = 3,005.").font = \
        font(9, italic=True, color="0F8B8D")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 46
    ws.freeze_panes = "B5"
    return {"takehome_row": th_row}

# ---- Multi-year & Year-2 --------------------------------------------------
def build_multiyear():
    ws = wb.create_sheet("Multi-year & Year-2")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Multi-year & Year-2", "Year-2 is the flagship lever")
    row = 4
    ws.cell(row=row, column=1, value="Multi-year discount").font = font(12, bold=True, color="0F8B8D"); row += 1
    header_row(ws, row, ZONES, "Metric"); row += 1
    ws.cell(row=row, column=1, value="Discount schedule").font = font(10, bold=True)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(row=row, column=1).border = BORDER; ws.cell(row=row, column=1).alignment = LEFT
    for i, z in enumerate(ZONES):
        style_cell(ws.cell(row=row, column=2 + i), MULTIYEAR_DISCOUNT[z])
    row += 2

    ws.cell(row=row, column=1, value="RAKEZ Biz One — cumulative price (AED)").font = font(11, bold=True, color="0F8B8D"); row += 1
    years = list(RAKEZ_MY_TABLE.keys())
    for j, y in enumerate(years):
        c = ws.cell(row=row, column=1 + j, value=y)
        c.font = font(10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="595959")
        c.alignment = CENTER; c.border = BORDER
    row += 1
    for j, y in enumerate(years):
        style_cell(ws.cell(row=row, column=1 + j), RAKEZ_MY_TABLE[y])
    row += 2

    ws.cell(row=row, column=1, value="Year-2 cost (1-visa, AED)").font = font(12, bold=True, color="0F8B8D"); row += 1
    header_row(ws, row, ZONES, "Metric"); row += 1
    ws.cell(row=row, column=1, value="Year-2 cost").font = font(10, bold=True)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
    ws.cell(row=row, column=1).border = BORDER; ws.cell(row=row, column=1).alignment = LEFT
    for i, z in enumerate(ZONES):
        style_cell(ws.cell(row=row, column=2 + i), YEAR2[z])
    row += 2

    ws.cell(row=row, column=1, value="KEY:").font = font(10, bold=True, color="C00000")
    ws.cell(row=row + 1, column=1, value=YEAR2_KEY).font = font(10, italic=True, color="595959")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 4, end_column=6)
    ws.cell(row=row + 1, column=1).alignment = LEFT

    ws.column_dimensions["A"].width = 26
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14

# ---- Sources --------------------------------------------------------------
def build_sources():
    ws = wb.create_sheet("Sources")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Sources", "Traceability — every figure tied to a source + confidence")
    row = 4
    headers = ["Area", "Zone", "Source", "Confidence", "Note"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + j, value=h)
        c.font = font(10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = CENTER; c.border = BORDER
    row += 1
    for area, zone, src, conf, note in SOURCES:
        ws.cell(row=row, column=1, value=area).font = font(10)
        ws.cell(row=row, column=2, value=zone).font = font(10)
        ws.cell(row=row, column=3, value=src).font = font(10)
        cc = ws.cell(row=row, column=4, value=LEGEND[conf][0])
        cc.font = font(10, bold=True, color=FONTCOL[conf]); cc.fill = FILL[conf]; cc.alignment = CENTER
        ws.cell(row=row, column=5, value=note).font = font(10)
        for j in range(5):
            ws.cell(row=row, column=1 + j).border = BORDER
        row += 1
    widths = [26, 10, 28, 16, 40]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.freeze_panes = "A5"

# ---- Assumptions & Gaps ---------------------------------------------------
def build_gaps():
    ws = wb.create_sheet("Assumptions & Gaps")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Assumptions & Gaps", "Placeholders to swap in — overwrite without touching the tool")
    row = 4
    headers = ["Item", "Current placeholder", "Confidence", "Owner", "Action to confirm"]
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + j, value=h)
        c.font = font(10, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = CENTER; c.border = BORDER
    row += 1
    for item, placeholder, conf, owner, action in GAPS:
        ws.cell(row=row, column=1, value=item).font = font(10, bold=True)
        ws.cell(row=row, column=2, value=placeholder).font = font(10)
        cc = ws.cell(row=row, column=3, value=LEGEND[conf][0])
        cc.font = font(10, bold=True, color=FONTCOL[conf]); cc.fill = FILL[conf]; cc.alignment = CENTER
        ws.cell(row=row, column=4, value=owner).font = font(10, bold=True)
        ws.cell(row=row, column=5, value=action).font = font(10)
        for j in range(5):
            ws.cell(row=row, column=1 + j).border = BORDER; ws.cell(row=row, column=1 + j).alignment = LEFT
        row += 1
    widths = [34, 32, 14, 10, 44]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.freeze_panes = "A5"

# ===========================================================================
# JSON builder (mirror of the workbook data)
# ===========================================================================
def cell_json(c):
    return {k: c[k] for k in ("kind", "value", "numeric", "display", "confidence", "note")}

def build_json():
    b2c = []
    for label, cells in B2C_ROWS:
        b2c.append({"component": label,
                    "zones": {z: cell_json(cells[z]) for z in ZONES}})
    base = {z: sum(r[1][z]["numeric"] for r in B2C_ROWS[0:3]) for z in ZONES}      # licence+reg+desk
    allin = {z: sum(r[1][z]["numeric"] for r in B2C_ROWS) for z in ZONES}
    # manual formula verification
    assert base == EXPECT_BASE, (base, EXPECT_BASE)
    assert allin == EXPECT_ALLIN, (allin, EXPECT_ALLIN)

    b2b = {key: {z: cell_json(B2B[key][z]) for z in ZONES} for _, key in B2B_ROW_ORDER}
    takehome = {z: round(tier_entry(z)["one_time"] * B2B["base_amount"][z]["numeric"]) for z in ZONES}
    assert takehome == EXPECT_TAKEHOME, (takehome, EXPECT_TAKEHOME)

    data = {
        "meta": {
            "title": "Competitor 360 — Master Data",
            "baseline": "Meydan",
            "zones": ZONES,
            "currency": "AED (VAT-incl)",
            "generated_from": "Competitor360_Master.xlsx",
            "note": "Single source of truth. Edit the XLSX, re-export. Do not hand-edit this file.",
        },
        "confidence_legend": {k: {"label": LEGEND[k][0], "meaning": LEGEND[k][1]}
                              for k in (CONFIRMED, ASSUMED, MISSING, NA)},
        "b2c": {
            "components": b2c,
            "base": base,
            "all_in": allin,
            "base_formula": "Licence + Registration + Shared Desk",
            "all_in_formula": "SUM of all components (incl/—/blank = 0)",
            "note": B2C_NOTE,
        },
        "b2b": {
            "rows": {key: b2b[key] for _, key in B2B_ROW_ORDER},
            "tiers": B2B_TIERS,
            "take_home": takehome,
            "take_home_formula": "entry one-time rate (tier rank 1) × base amount",
            "level_note": B2B_LEVEL_NOTE,
            "level_descriptions": B2B_LEVEL_DESCRIPTIONS,
        },
        "multi_year": {
            "discount": {z: cell_json(MULTIYEAR_DISCOUNT[z]) for z in ZONES},
            "rakez_biz_one": {y: cell_json(RAKEZ_MY_TABLE[y]) for y in RAKEZ_MY_TABLE},
            "year_2": {z: cell_json(YEAR2[z]) for z in ZONES},
            "year_2_key": YEAR2_KEY,
        },
        "sources": [{"area": a, "zone": z, "source": s, "confidence": c, "note": n}
                    for a, z, s, c, n in SOURCES],
        "assumptions_gaps": [{"item": i, "placeholder": p, "confidence": c, "owner": o, "action": ac}
                             for i, p, c, o, ac in GAPS],
    }
    return data

# ===========================================================================
# Run
# ===========================================================================
build_readme()
b2c_meta = build_b2c()
b2b_meta = build_b2b()
build_multiyear()
build_sources()
build_gaps()

wb.save(r"D:\COMPETITOR360_V2\Competitor360_Master.xlsx")

data = build_json()
with open(r"D:\COMPETITOR360_V2\master-data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("OK — XLSX + JSON written.")
print("BASE :", data["b2c"]["base"])
print("ALLIN:", data["b2c"]["all_in"])
print("TAKEHOME:", data["b2b"]["take_home"])
